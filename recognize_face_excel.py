import cv2
import pickle
import numpy as np
from keras_facenet import FaceNet
from datetime import datetime, date
import time
from openpyxl import Workbook, load_workbook
import os
import winsound  # 🔊 ใช้สร้างเสียง beep (Windows เท่านั้น)
import requests  # 📲 สำหรับส่งข้อความแจ้งเตือนผ่าน Line Notify

# ============================
# 🔥 โหลดโมเดล FaceNet
# ใช้แปลง “ใบหน้า → vector (embedding)”
# ============================
embedder = FaceNet()

# ============================
# 🔥 โหลดโมเดล SVM
# ใช้ทำนายว่า embedding นี้เป็น “ใคร”
# ============================
model = pickle.load(open("facenet_svm.pkl", "rb"))

# ============================
# 🔥 โหลด Label Encoder
# ใช้แปลงเลข → ชื่อคน
# ============================
encoder = pickle.load(open("label_encoder.pkl", "rb"))

# ============================
# 🔥 โหลด Haar Cascade
# ใช้ตรวจจับตำแหน่งใบหน้าในภาพ
# ============================
face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades +
    "haarcascade_frontalface_default.xml"
)

# ============================
# 🔥 ตั้งค่าไฟล์ Excel สำหรับบันทึกการเข้าเรียน
# ============================
excel_file = "attendance.xlsx"

# ถ้ายังไม่มีไฟล์ → สร้างใหม่
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Date", "Time"])  # header
    wb.save(excel_file)

# โหลดไฟล์ Excel
wb = load_workbook(excel_file)
ws = wb.active

# ============================
# 🔥 ตั้งค่า Line Messaging API
# ============================
LINE_CHANNEL_ACCESS_TOKEN = "vo5Zj/gXrwfr9YL+y6Rc131EBmaXXabL7onkIjLevS3vgVXDmpzXlcaZffj/8QVhH3VOBwTJ0EnLO0TAPA/vm2AH+6+8WEVSYXvzxLRt+3ow7vgvX3lPPtQP5AutEDYjxaspTziIEGOLH1StANairwdB04t89/1O/w1cDnyilFU="  # 🟢 Channel Access Token
LINE_USER_ID = "U252bb0d0b12fee7563f7ae0b8e601dd7"              # 🟢 User ID ของคุณ

def send_line_notify(message, image_path=None):
    if not LINE_CHANNEL_ACCESS_TOKEN or LINE_CHANNEL_ACCESS_TOKEN == "ใส่ Channel Access Token ที่นี่":
        print("[LINE] ข้ามการส่งแจ้งเตือน: ยังไม่ได้ระบุ LINE_CHANNEL_ACCESS_TOKEN")
        return False

    url = "https://api.line.me/v2/bot/message/push"
    headers = {
        "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }

    # ส่งข้อความ text
    payload = {
        "to": LINE_USER_ID,
        "messages": [
            {
                "type": "text",
                "text": message
            }
        ]
    }

    try:
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 200:
            print(f"[LINE NOTIFY] ส่งการแจ้งเตือนสำเร็จ: {message}")
            return True
        else:
            print(f"[LINE NOTIFY] ส่งการแจ้งเตือนไม่สำเร็จ รหัสสถานะ: {response.status_code}, ผลลัพธ์: {response.text}")
            return False
    except Exception as e:
        print(f"[LINE NOTIFY] เกิดข้อผิดพลาดในการส่งข้อมูล: {e}")
        return False

# ============================
# 🔥 โหลดข้อมูลจาก Excel เพื่อป้องกันการเช็คชื่อซ้ำในวันเดียวกัน (กรณีเปิดโปรแกรมใหม่)
# ============================
checked_today = set()

# โหลดประวัติที่มีอยู่แล้วใน Excel มาใส่ใน checked_today
if ws.max_row >= 2:
    for row in range(2, ws.max_row + 1):
        name_val = ws.cell(row=row, column=1).value
        date_val = ws.cell(row=row, column=2).value
        if name_val and date_val:
            if isinstance(date_val, (date, datetime)):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val).strip()
            checked_today.add(f"{name_val}_{date_str}")

print(f"Loaded {len(checked_today)} existing check-ins to prevent duplicates.")

# วันที่ปัจจุบัน
today_date = date.today()
today_str = today_date.strftime("%Y-%m-%d")

# ============================
# 🔊 ระบบเสียง (กันเสียงรัว)
# ============================
last_beep_time = 0          # เวลาที่ beep ล่าสุด
BEEP_INTERVAL = 3           # เว้น 3 วินาทีค่อย beep อีกครั้ง

# ============================
# 📷 เปิดกล้อง webcam
# ============================
cap = cv2.VideoCapture(0)

while True:

    # อ่านภาพจากกล้อง
    ret, frame = cap.read()
    if not ret:
        break

    # แปลงภาพเป็น grayscale เพื่อใช้ตรวจจับหน้า
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    # ============================
    # 🔥 ตรวจจับใบหน้าในภาพ
    # ============================
    faces = face_cascade.detectMultiScale(gray, 1.2, 5)

    for (x, y, w, h) in faces:

        # ============================
        # crop เฉพาะใบหน้า
        # ============================
        face = frame[y:y+h, x:x+w]

        if face.size == 0:
            continue

        # resize ให้ตรงกับ FaceNet (160x160)
        face = cv2.resize(face, (160, 160))

        # แปลง BGR → RGB (FaceNet ใช้ RGB)
        rgb_face = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)

        # ============================
        # 🔥 แปลงใบหน้าเป็น embedding vector
        # ============================
        embedding = embedder.embeddings([rgb_face])[0]

        # normalize vector เพื่อให้แม่นขึ้น
        embedding = embedding / np.linalg.norm(embedding)

        # reshape ให้ model ใช้ได้
        embedding = np.expand_dims(embedding, axis=0)

        # ============================
        # 🔥 ทำนายชื่อด้วย SVM
        # ============================
        pred = model.predict(embedding)[0]

        # ความน่าจะเป็นของผลลัพธ์ทั้งหมด
        proba = model.predict_proba(embedding)[0]

        # confidence สูงสุด
        confidence = np.max(proba)

        # แปลงเลข label → ชื่อคน
        name = encoder.inverse_transform([pred])[0]

        # ============================
        # 🔥 ถ้าความมั่นใจต่ำ → Unknown
        # ============================
        if confidence < 0.90:
            name = "Unknown"

        # ============================
        # 🔥 ระบบเช็คชื่อเข้าเรียน
        # ============================
        if name != "Unknown":

            # วันที่ปัจจุบัน (คำนวณแบบ dynamic ป้องกันรันข้ามคืนแล้ววันที่ไม่อัปเดต)
            current_date_str = date.today().strftime("%Y-%m-%d")

            # สร้าง key กันซ้ำ (ชื่อ + วันที่)
            key = f"{name}_{current_date_str}"

            # ถ้ายังไม่เคยเช็ควันนี้
            if key not in checked_today:

                now = datetime.now()
                time_str = now.strftime("%H:%M:%S")

                # บันทึกลง Excel
                ws.append([name, current_date_str, time_str])
                wb.save(excel_file)

                # เพิ่มเข้า set กันซ้ำ
                checked_today.add(key)

                print(f"[CHECK-IN] {name} at {time_str}")

                # ============================
                # 📸 เซฟภาพใบหน้าที่สแกนสำเร็จ เพื่อส่งแจ้งเตือน Line
                # ============================
                temp_img_path = "temp_scan.jpg"
                
                # วาดกรอบรอบหน้าและชื่อบน frame ที่จะบันทึกรูป
                scan_frame = frame.copy()
                color = (0, 255, 0)
                cv2.rectangle(scan_frame, (x, y), (x+w, y+h), color, 2)
                cv2.putText(
                    scan_frame,
                    f"{name} {confidence:.2f}",
                    (x, y-10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.8,
                    color,
                    2
                )
                
                cv2.imwrite(temp_img_path, scan_frame)
                
                # ส่งข้อความแจ้งเตือนผ่าน Line Messaging API (format เหมือน Excel)
                line_sep = chr(9472) * 25
                line_msg = (
                    "📋 ระบบเช็คชื่อเข้าเรียน\n" +
                    line_sep + "\n" +
                    "Name  : " + name + "\n" +
                    "Date  : " + current_date_str + "\n" +
                    "Time  : " + time_str + "\n" +
                    line_sep + "\n" +
                    "✅ เช็คอินสำเร็จ"
                )
                send_line_notify(line_msg, temp_img_path)
                
                # ลบไฟล์ภาพชั่วคราว
                try:
                    if os.path.exists(temp_img_path):
                        os.remove(temp_img_path)
                except Exception as e:
                    pass

                # ============================
                # 🔊 เล่นเสียงแจ้งเตือน
                # ============================
                current_time = time.time()

                # กันเสียงรัว
                if current_time - last_beep_time > BEEP_INTERVAL:
                    winsound.Beep(1200, 200)  # เสียง beep
                    last_beep_time = current_time

        # ============================
        # 🔥 วาดกรอบใบหน้า
        # ============================
        color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)

        cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)

        cv2.putText(
            frame,
            f"{name} {confidence:.2f}",
            (x, y-10),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.8,
            color,
            2
        )

    # ============================
    # 🔥 แสดงผลหน้าจอ
    # ============================
    cv2.imshow("Attendance System", frame)

    # กด q เพื่อออกจากโปรแกรม
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# ============================
# 🔥 ปิดกล้อง + เซฟไฟล์
# ============================
cap.release()
wb.save(excel_file)
cv2.destroyAllWindows()